import streamlit as st
import pandas as pd
import psycopg2
import psycopg2.extras
import io
import os
from datetime import datetime

# ── Config ────────────────────────────────────────────────────
st.set_page_config(
    page_title="HAI-Q · Painel de Acompanhamento",
    page_icon="📊",
    layout="wide",
)

BASE_URL     = os.environ.get("HAIQ_BASE_URL", "https://assessmenthaiq-production.up.railway.app")
DATABASE_URL = os.environ.get("DATABASE_URL", "")

# ── Estilos ───────────────────────────────────────────────────
st.markdown("""
<style>
    .block-container { padding-top: 1.5rem; }
    .metric-card {
        background: #f8f9fb;
        border: 1px solid #e0e7ef;
        border-radius: 10px;
        padding: 1rem 1.25rem;
        text-align: center;
    }
    .metric-num  { font-size: 2rem; font-weight: 600; margin: 0; }
    .metric-lbl  { font-size: 0.8rem; color: #666; margin: 0; }
    .status-ok   { color: #0F6E56; font-weight: 600; }
    .status-wait { color: #185FA5; font-weight: 600; }
    .status-warn { color: #854F0B; font-weight: 600; }
    .status-no   { color: #D85A30; font-weight: 600; }
</style>
""", unsafe_allow_html=True)

# ── Header ────────────────────────────────────────────────────
col_logo, col_title = st.columns([1, 8])
with col_logo:
    st.markdown("## HAI-Q")
with col_title:
    st.markdown("### Painel de Acompanhamento de Candidatos")

st.divider()

# ── Conexão ───────────────────────────────────────────────────
@st.cache_resource
def get_conn():
    url = DATABASE_URL
    if not url:
        return None
    if "sslmode" not in url:
        sep = "&" if "?" in url else "?"
        url = url + sep + "sslmode=require"
    return psycopg2.connect(url)

def query_emails(emails: list[str]) -> pd.DataFrame:
    conn = get_conn()
    if not conn:
        st.error("DATABASE_URL não configurada nas variáveis de ambiente.")
        return pd.DataFrame()

    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    rows = []

    for email in emails:
        email = email.strip().lower()
        if not email:
            continue
        try:
            cur.execute("""
                SELECT
                    s.id                    AS session_id,
                    s."candidateName"       AS nome,
                    s."candidateEmail"      AS email,
                    s."caseTitle"           AS case_titulo,
                    s."startedAt"           AS iniciou_em,
                    s."completedAt"         AS finalizou_em,
                    s."durationSeconds"     AS duracao_segundos,
                    e."haiqScore"           AS haiq_score,
                    e."d1Score"             AS d1,
                    e."d2Score"             AS d2,
                    e."d3Score"             AS d3,
                    e."d4Score"             AS d4,
                    e."agencyClass"         AS agencia,
                    e."evaluatedAt"         AS avaliado_em
                FROM "Session" s
                LEFT JOIN "Evaluation" e ON e."sessionId" = s.id
                WHERE LOWER(s."candidateEmail") = %s
                ORDER BY s."startedAt" ASC
            """, (email,))
            found = cur.fetchall()
        except Exception as ex:
            conn.rollback()
            st.warning(f"Erro ao buscar {email}: {ex}")
            continue

        if not found:
            rows.append({
                "Email": email, "Nome": "—", "Tentativa": "—",
                "Status": "❌ Não iniciou",
                "Case": "—", "Iniciou em": "—", "Finalizou em": "—",
                "Duração (min)": "—", "HAI-Q": "—",
                "D1": "—", "D2": "—", "D3": "—", "D4": "—",
                "Agência": "—",
                "session_id": None,
                "finalizado": False,
            })
        else:
            total = len(found)
            for i, r in enumerate(found, 1):
                sid       = r["session_id"]
                finalizou = r["finalizou_em"]
                haiq      = r["haiq_score"]
                duracao   = r["duracao_segundos"]

                tentativa = "Única" if total == 1 else f"{i} de {total}"

                if not finalizou:
                    status = "⚠️ Iniciou mas não finalizou"
                elif not haiq:
                    status = "🔄 Finalizado — aguardando avaliação"
                else:
                    status = "✅ Concluído e avaliado"

                rows.append({
                    "Email":          r["email"],
                    "Nome":           r["nome"],
                    "Tentativa":      tentativa,
                    "Status":         status,
                    "Case":           r["case_titulo"] or "—",
                    "Iniciou em":     r["iniciou_em"].strftime("%d/%m/%Y %H:%M") if r["iniciou_em"] else "—",
                    "Finalizou em":   finalizou.strftime("%d/%m/%Y %H:%M") if finalizou else "—",
                    "Duração (min)":  round(duracao / 60, 1) if duracao else "—",
                    "HAI-Q":          round(haiq, 2) if haiq else "—",
                    "D1":             round(r["d1"], 1) if r["d1"] else "—",
                    "D2":             round(r["d2"], 1) if r["d2"] else "—",
                    "D3":             round(r["d3"], 1) if r["d3"] else "—",
                    "D4":             round(r["d4"], 1) if r["d4"] else "—",
                    "Agência":        r["agencia"] or "—",
                    "session_id":     sid,
                    "finalizado":     bool(finalizou),
                })

    cur.close()
    return pd.DataFrame(rows)

# ── Entrada de e-mails ────────────────────────────────────────
st.markdown("#### Como deseja informar os e-mails?")
tab_upload, tab_digitar = st.tabs(["📂 Upload de Excel ou CSV", "✏️ Digitar manualmente"])

emails_input: list[str] = []

with tab_upload:
    uploaded = st.file_uploader(
        "Envie um arquivo Excel (.xlsx) ou CSV com uma coluna de e-mails",
        type=["xlsx", "csv"],
    )
    col_name = st.text_input("Nome da coluna de e-mails no arquivo", value="email")

    if uploaded:
        try:
            if uploaded.name.endswith(".csv"):
                df_up = pd.read_csv(uploaded)
            else:
                df_up = pd.read_excel(uploaded)

            if col_name not in df_up.columns:
                st.error(f"Coluna '{col_name}' não encontrada. Colunas disponíveis: {list(df_up.columns)}")
            else:
                emails_input = df_up[col_name].dropna().str.strip().str.lower().unique().tolist()
                st.success(f"{len(emails_input)} e-mail(s) carregado(s)")
                with st.expander("Ver e-mails carregados"):
                    st.write(emails_input)
        except Exception as e:
            st.error(f"Erro ao ler arquivo: {e}")

with tab_digitar:
    texto = st.text_area(
        "Digite os e-mails, um por linha",
        placeholder="joao@empresa.com\nmaria@empresa.com\npedro@empresa.com",
        height=150,
    )
    if texto.strip():
        emails_input = [e.strip().lower() for e in texto.strip().splitlines() if e.strip()]
        st.info(f"{len(emails_input)} e-mail(s) informado(s)")

# ── Buscar ────────────────────────────────────────────────────
st.divider()

if emails_input:
    if st.button("🔍 Buscar status", type="primary", use_container_width=True):
        with st.spinner("Consultando o banco..."):
            df = query_emails(emails_input)
            st.session_state["df_result"] = df
            st.session_state["searched"]  = True

# Mostrar resultado se já buscou
if st.session_state.get("searched") and "df_result" in st.session_state:
    df: pd.DataFrame = st.session_state["df_result"]

    if df.empty:
        st.warning("Nenhum resultado encontrado.")
    else:
        # ── Métricas resumo ───────────────────────────────────
        total       = df["Email"].nunique()
        nao_iniciou = df[df["Status"].str.contains("Não iniciou",   na=False)]["Email"].nunique()
        nao_fin     = df[df["Status"].str.contains("não finalizou", na=False)]["Email"].nunique()
        aguardando  = df[df["Status"].str.contains("aguardando",    na=False)]["Email"].nunique()
        avaliado    = df[df["Status"].str.contains("avaliado",      na=False)]["Email"].nunique()
        multiplas   = df[df["Tentativa"].str.contains(" de ",       na=False)]["Email"].nunique()

        c1, c2, c3, c4, c5, c6 = st.columns(6)
        for col, num, lbl, cor in [
            (c1, total,       "Total",                  "#042C53"),
            (c2, avaliado,    "✅ Avaliados",            "#0F6E56"),
            (c3, aguardando,  "🔄 Aguardando avaliação", "#185FA5"),
            (c4, nao_fin,     "⚠️ Não finalizou",        "#854F0B"),
            (c5, nao_iniciou, "❌ Não iniciou",          "#D85A30"),
            (c6, multiplas,   "🔁 Múltiplas tentativas", "#534AB7"),
        ]:
            col.markdown(
                f'<div class="metric-card">'
                f'<p class="metric-num" style="color:{cor}">{num}</p>'
                f'<p class="metric-lbl">{lbl}</p>'
                f'</div>',
                unsafe_allow_html=True,
            )

        st.markdown("")

        # ── Tabela principal ──────────────────────────────────
        df_display = df.drop(columns=["session_id", "finalizado"], errors="ignore").copy()

        # Adicionar colunas de link como texto
        def make_link(row: pd.Series, tipo: str) -> str:
            if not row.get("finalizado"):
                return "—"
            sid = row.get("session_id")
            if not sid:
                return "—"
            if tipo == "candidato":
                return f"{BASE_URL}/devolutiva/{sid}"
            return f"{BASE_URL}/admin/devolutiva/{sid}"

        df["link_cand"] = df.apply(lambda r: make_link(r, "candidato"), axis=1)
        df["link_rh"]   = df.apply(lambda r: make_link(r, "rh"), axis=1)

        # Tabela exibida sem colunas internas
        cols_show = ["Email", "Nome", "Tentativa", "Status", "Case",
                     "Iniciou em", "Finalizou em", "Duração (min)",
                     "HAI-Q", "D1", "D2", "D3", "D4", "Agência"]
        st.dataframe(
            df[cols_show],
            use_container_width=True,
            hide_index=True,
            column_config={
                "HAI-Q": st.column_config.NumberColumn(format="%.2f"),
                "D1":    st.column_config.NumberColumn(format="%.1f"),
                "D2":    st.column_config.NumberColumn(format="%.1f"),
                "D3":    st.column_config.NumberColumn(format="%.1f"),
                "D4":    st.column_config.NumberColumn(format="%.1f"),
            }
        )

        # ── Links de devolutiva ───────────────────────────────
        finalizados = df[df["finalizado"] == True].copy()
        if not finalizados.empty:
            st.markdown("#### 🔗 Links de Devolutiva")
            for _, row in finalizados.iterrows():
                sid   = row["session_id"]
                nome  = row["Nome"]
                tent  = row["Tentativa"]
                email = row["Email"]
                label = f"**{nome}** ({email})" + (f" — tentativa {tent}" if tent != "Única" else "")
                url_cand = f"{BASE_URL}/devolutiva/{sid}"
                url_rh   = f"{BASE_URL}/admin/devolutiva/{sid}"

                st.markdown(label)
                c_link, c_rh = st.columns(2)
                c_link.markdown(
                    "👤 **Devolutiva Candidato**  \n"
                    f'<a href="{url_cand}" target="_blank">{url_cand}</a>',
                    unsafe_allow_html=True,
                )
                c_rh.markdown(
                    "📋 **Devolutiva RH**  \n"
                    f'<a href="{url_rh}" target="_blank">{url_rh}</a>',
                    unsafe_allow_html=True,
                )
                st.divider()

        st.divider()

        # ── Export Excel ──────────────────────────────────────
        df_export = df[cols_show + ["link_cand", "link_rh"]].rename(columns={
            "link_cand": "Link Devolutiva (Candidato)",
            "link_rh":   "Link Devolutiva (RH)",
        })

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df_export.to_excel(writer, index=False, sheet_name="Status")
            ws = writer.sheets["Status"]

            from openpyxl.styles import PatternFill, Font
            from openpyxl.utils import get_column_letter

            widths = [30,25,12,35,35,18,18,14,10,10,10,10,10,18,50,50]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

            fills = {
                "✅": PatternFill(start_color="E1F5EE", end_color="E1F5EE", fill_type="solid"),
                "🔄": PatternFill(start_color="E6F1FB", end_color="E6F1FB", fill_type="solid"),
                "⚠️": PatternFill(start_color="FAEEDA", end_color="FAEEDA", fill_type="solid"),
                "❌": PatternFill(start_color="FAECE7", end_color="FAECE7", fill_type="solid"),
            }
            fill_repeat = PatternFill(start_color="F1EFE8", end_color="F1EFE8", fill_type="solid")

            for row_idx, row_data in enumerate(df_export.itertuples(), start=2):
                status    = str(row_data.Status)
                tentativa = str(row_data.Tentativa)
                is_repeat = " de " in tentativa and not tentativa.startswith("1 de")

                for col_idx in range(1, len(df_export.columns) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if is_repeat:
                        cell.fill = fill_repeat
                    else:
                        for emoji, fill in fills.items():
                            if emoji in status:
                                cell.fill = fill
                                break

                # Links clicáveis
                lc = getattr(row_data, "Link_Devolutiva__Candidato_", "—")
                lr = getattr(row_data, "Link_Devolutiva__RH_", "—")
                if lc and lc != "—":
                    cell = ws.cell(row=row_idx, column=15)
                    cell.hyperlink = lc
                    cell.value     = "Abrir devolutiva candidato"
                    cell.font      = Font(color="185FA5", underline="single")
                if lr and lr != "—":
                    cell = ws.cell(row=row_idx, column=16)
                    cell.hyperlink = lr
                    cell.value     = "Abrir devolutiva RH"
                    cell.font      = Font(color="0F6E56", underline="single")

            for col_idx in range(1, len(df_export.columns) + 1):
                ws.cell(row=1, column=col_idx).font = Font(bold=True)

        buf.seek(0)
        st.download_button(
            label="📥 Baixar Excel com todos os resultados",
            data=buf,
            file_name=f"haiq_status_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

else:
    st.info("Informe os e-mails na aba acima e clique em **Buscar status**.")
